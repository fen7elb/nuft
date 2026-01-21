import openpyxl
import os

file_path = 'data/Довідник НУХТ.xlsx'

if not os.path.exists(file_path):
    print(f"erroe: файл не знайдено за адресою: {file_path}")
else:
    print(f"файл є.")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    print("\nсписок вкладок у файлі:")
    print(wb.sheetnames)

    print("\nтест звчитування (перші 3 рядки з кожної вкладки)")
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\nвкладка: [{sheet_name}]")
        for i, row in enumerate(sheet.iter_rows(max_row=3, values_only=True)):
            print(f"  Рядок {i+1}: {row[:3]}")