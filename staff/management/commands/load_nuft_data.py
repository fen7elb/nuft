import openpyxl
import os
import re
from django.core.management.base import BaseCommand
from staff.models import Department, Employee, Position, TypeDepartment
from django.conf import settings

class Command(BaseCommand):
    help = 'імпорт даних нухт: збереження ID + розумні абревіатури'

    def handle(self, *args, **kwargs):
        file_path = settings.BASE_DIR / 'data' / 'Довідник НУХТ.xlsx'
        
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f'файл {file_path} не знайдено!'))
            return

        self.stdout.write(self.style.WARNING('Починаємо розумний імпорт (оновлення)...'))
        
        # Списки "живих" ID, щоб знати, кого НЕ видаляти
        active_department_ids = []
        active_employee_ids = []
        
        self.stdout.write(self.style.SUCCESS(f'відкриваємо Excel...'))
        workbook = openpyxl.load_workbook(file_path, data_only=True)

        sheets_map = {
            'Ректорат': 'Ректорат',
            'Інститути': 'Інститут',
            'Деканати': 'Факультет',
            'Кафедри': 'Кафедра',
            'Відділи': 'Відділ'
        }

        for sheet_name, type_name in sheets_map.items():
            if sheet_name not in workbook.sheetnames:
                continue

            # Тип підрозділу створюємо або отримуємо
            type_dep, _ = TypeDepartment.objects.get_or_create(name_type=type_name)
            sheet = workbook[sheet_name]
            
            self.stdout.write(self.style.SUCCESS(f'обробка вкладки: {sheet_name} '))
            current_department = None

            for row in sheet.iter_rows(min_row=2, values_only=True):
                val0 = str(row[0]).strip() if row[0] else ""
                val1 = str(row[1]).strip() if row[1] else ""
                val2 = str(row[2]).strip() if row[2] else ""
                val3 = str(row[3]).strip() if row[3] else ""
                val4 = str(row[4]).strip() if len(row) > 4 and row[4] else ""

                if not val0 and not val1:
                    continue 

                # --- КАФЕДРА/ВІДДІЛ ---
                if val0 and not val1:
                    raw_text = val0
                    
                    found_email = re.search(r'[\w\.-]+@[\w\.-]+\.\w+', raw_text)
                    dept_mail = val2 
                    if found_email:
                        email_str = found_email.group(0)
                        if not dept_mail: dept_mail = email_str
                        raw_text = raw_text.replace(email_str, '')

                    # Чистка назви
                    clean_name = re.sub(r'^\(\d+\)\s*', '', raw_text)
                    clean_name = clean_name.replace('\n', ' ').strip()
                    clean_name = re.sub(r'\s+', ' ', clean_name)
                    
                    if clean_name.isupper():
                        clean_name = clean_name.capitalize()

                    # Абревіатури для Інститутів/Факультетів
                    short_val = None
                    if type_name in ['Інститут', 'Факультет']:
                        match_abbr = re.search(r'\((.*?)\)', clean_name)
                        if match_abbr:
                            short_val = match_abbr.group(1)[:50]
                        else:
                            short_val = clean_name[:50]

                    # Оновлюємо або створюємо (щоб ID не мінявся)
                    current_department, created = Department.objects.update_or_create(
                        name=clean_name,
                        defaults={
                            'short_name': short_val,
                            'mail': dept_mail,
                            'type_dep': type_dep
                        }
                    )
                    active_department_ids.append(current_department.pk)
                    continue

                # --- СПІВРОБІТНИК ---
                if val1 and current_department:
                    full_name_clean = val1.replace('\n', ' ').strip().title()
                    position_clean = val0.replace('\n', ' ').strip().capitalize()
                    phone_clean = val3.replace('\n', ', ').strip()[:20]
                    office_clean = val4.replace('\n', ' ').strip()[:20]

                    position_obj, _ = Position.objects.get_or_create(title=position_clean)

                    # Оновлюємо або створюємо співробітника
                    employee, created = Employee.objects.update_or_create(
                        full_name=full_name_clean,
                        department=current_department,
                        defaults={
                            'phone_number': phone_clean,
                            'mail': val2,
                            'office': office_clean,
                            'position': position_obj
                        }
                    )
                    active_employee_ids.append(employee.pk)

        # Видаляємо тільки тих, кого НЕМАЄ в новому файлі
        deleted_emps, _ = Employee.objects.exclude(pk__in=active_employee_ids).delete()
        if deleted_emps:
            self.stdout.write(self.style.WARNING(f'Видалено звільнених: {deleted_emps}'))

        deleted_depts, _ = Department.objects.exclude(pk__in=active_department_ids).delete()
        if deleted_depts:
            self.stdout.write(self.style.WARNING(f'Видалено закритих: {deleted_depts}'))

        self.stdout.write(self.style.SUCCESS('Імпорт завершено! База актуальна.'))